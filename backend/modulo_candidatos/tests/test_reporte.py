# ==============================================================
# modulo_candidatos / tests/test_reporte.py
# ==============================================================

import os
import sys
import io
from datetime import datetime, timezone

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import openpyxl
import pytest

from tests.conftest import client
from tests.test_solicitudes import HEADERS_CANDIDATO, HEADERS_GESTOR, SOLICITUD_VALIDA
from app.services import reporte_service

pytestmark = pytest.mark.usefixtures("mock_vacantes")


class _FakeSolicitud:
    pass


def _solicitud_fake(**kwargs):
    s = _FakeSolicitud()
    s.datos_personales = kwargs.get("datos_personales", {"cedula": "123", "nombreCompleto": "Ana Prueba", "correo": "ana@example.com", "celular": "3000000000"})
    s.evaluacion = kwargs.get("evaluacion", None)
    return s


VACANTE_INFO = {
    "proceso_no": "2026-1", "cargo": "Analista de Prueba", "salario": "3.000.000",
    "fecha_apertura": "2026-01-01", "fecha_cierre": "2026-02-01",
    "criterios": {"nivel_educativo_min": "Universitario", "graduado_requerido": True, "experiencia_min_anios": 2},
}


def test_generar_reporte_produce_un_xlsx_con_las_2_hojas():
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [_solicitud_fake()])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    assert wb.sheetnames == ["Aspirantes", "Resultados_reclutamiento"]


def test_reporte_encabezado_hoja_aspirantes():
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [_solicitud_fake()])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb["Aspirantes"]
    assert ws["D4"].value == "Analista de Prueba"
    assert ws["D5"].value == "2026-1"
    assert ws["D6"].value == "2026-01-01"
    assert ws["D7"].value == "2026-02-01"
    assert "Universitario" in ws["D8"].value
    assert "2" in ws["D9"].value


def test_reporte_encabezado_hoja_resultados_reclutamiento_concatena_con_la_etiqueta():
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [_solicitud_fake()])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb["Resultados_reclutamiento"]
    assert ws["D5"].value == "PROCESO DE SELECCIÓN Nº: 2026-1"
    assert ws["A6"].value == "CARGO: Analista de Prueba"
    assert ws["A7"].value == "SALARIO BASICO: 3.000.000"


def test_reporte_datos_del_candidato_en_hoja_aspirantes():
    s = _solicitud_fake(datos_personales={"cedula": "999", "nombreCompleto": "Pedro Real", "correo": "pedro@example.com", "celular": "3001112233"})
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [s])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb["Aspirantes"]
    assert ws["A13"].value == "999"
    assert ws["B13"].value == "Pedro Real"
    assert ws["L13"].value == "pedro@example.com"
    assert ws["M13"].value == "3001112233"


def test_reporte_una_fila_por_solicitud_en_orden_hoja_aspirantes():
    solicitudes = [
        _solicitud_fake(datos_personales={"cedula": "111", "nombreCompleto": "Primero"}),
        _solicitud_fake(datos_personales={"cedula": "222", "nombreCompleto": "Segundo"}),
    ]
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, solicitudes)
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb["Aspirantes"]
    assert ws["A13"].value == "111"
    assert ws["A14"].value == "222"
    assert ws["A15"].value is None


def test_reporte_una_fila_por_solicitud_en_orden_hoja_resultados():
    solicitudes = [
        _solicitud_fake(datos_personales={"cedula": "111", "nombreCompleto": "Primero"}),
        _solicitud_fake(datos_personales={"cedula": "222", "nombreCompleto": "Segundo"}),
    ]
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, solicitudes)
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb["Resultados_reclutamiento"]
    assert ws["A16"].value == "111"
    assert ws["A17"].value == "222"


def test_reporte_marca_x_en_columna_si_cuando_categoria_cumple():
    s = _solicitud_fake(evaluacion={"cumple": True, "motivos": [], "detalle": {
        "estudios": {"cumple": True, "motivo": None},
        "conocimientos": {"cumple": None, "motivo": None},
        "experiencia": {"cumple": True, "motivo": None},
    }})
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [s])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    for hoja, fila in [("Aspirantes", 13), ("Resultados_reclutamiento", 16)]:
        ws = wb[hoja]
        assert ws[f"C{fila}"].value == "X"  # Estudios - SI
        assert ws[f"D{fila}"].value is None
        assert ws[f"G{fila}"].value == "X"  # Experiencia - SI
        assert ws[f"E{fila}"].value is None  # Conocimientos: sin criterio, blanco


def test_reporte_marca_x_en_columna_no_cuando_categoria_no_cumple():
    s = _solicitud_fake(evaluacion={"cumple": False, "motivos": ["algo"], "detalle": {
        "estudios": {"cumple": False, "motivo": "No cumple nivel"},
        "conocimientos": {"cumple": None, "motivo": None},
        "experiencia": {"cumple": None, "motivo": None},
    }})
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [s])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    for hoja, fila in [("Aspirantes", 13), ("Resultados_reclutamiento", 16)]:
        ws = wb[hoja]
        assert ws[f"D{fila}"].value == "X"  # Estudios - NO
        assert ws[f"C{fila}"].value is None


def test_reporte_no_marca_decision_automaticamente():
    s = _solicitud_fake(evaluacion={"cumple": True, "motivos": [], "detalle": {
        "estudios": {"cumple": True, "motivo": None},
        "conocimientos": {"cumple": True, "motivo": None},
        "experiencia": {"cumple": True, "motivo": None},
    }})
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [s])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    for hoja, fila in [("Aspirantes", 13), ("Resultados_reclutamiento", 16)]:
        ws = wb[hoja]
        assert ws[f"I{fila}"].value is None
        assert ws[f"J{fila}"].value is None


def test_reporte_sin_evaluacion_queda_todo_en_blanco():
    s = _solicitud_fake(evaluacion=None)
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, [s])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb["Aspirantes"]
    for col in ["C", "D", "E", "F", "G", "H"]:
        assert ws[f"{col}13"].value is None


def test_reporte_sin_criterios_configurados_describe_sin_requisito():
    vacante_sin_criterios = {**VACANTE_INFO, "criterios": {}}
    contenido = reporte_service.generar_reporte_vacante(vacante_sin_criterios, [_solicitud_fake()])
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb["Aspirantes"]
    assert "Sin requisito específico" in ws["D8"].value
    assert "Sin requisito específico" in ws["D9"].value
    assert "Sin requisito específico" in ws["D10"].value


def test_reporte_no_tiene_errores_de_formula_tras_recalcular(tmp_path):
    import subprocess
    import json

    solicitudes = [_solicitud_fake(datos_personales={"cedula": str(i), "nombreCompleto": f"Persona {i}"}) for i in range(5)]
    contenido = reporte_service.generar_reporte_vacante(VACANTE_INFO, solicitudes)
    ruta = tmp_path / "reporte.xlsx"
    ruta.write_bytes(contenido)

    resultado = subprocess.run(
        ["python3", "/mnt/skills/public/xlsx/scripts/recalc.py", str(ruta)],
        capture_output=True, text=True, timeout=60,
    )
    salida = json.loads(resultado.stdout)
    assert salida["status"] == "success"
    assert salida["total_errors"] == 0


def test_descargar_reporte_requiere_rol_gestion():
    r = client.get("/api/v1/solicitudes/admin/reporte/vac-1", headers=HEADERS_CANDIDATO)
    assert r.status_code == 403


def test_descargar_reporte_vacante_inexistente_da_404():
    r = client.get("/api/v1/solicitudes/admin/reporte/no-existe", headers=HEADERS_GESTOR)
    assert r.status_code == 404


def test_descargar_reporte_devuelve_un_xlsx_valido():
    client.post("/api/v1/solicitudes", json=SOLICITUD_VALIDA, headers=HEADERS_CANDIDATO)

    r = client.get("/api/v1/solicitudes/admin/reporte/vac-1", headers=HEADERS_GESTOR)
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in r.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Aspirantes"]
    assert ws["A13"].value == SOLICITUD_VALIDA["datos_personales"]["cedula"]
